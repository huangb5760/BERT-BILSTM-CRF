from openpyxl import load_workbook
import json
import os
import re
from predict import  Predictor

class ProcessWineData:
	def __init__(self):
		self.data_path = "./data/steel/"
		self.train_file_ori = self.data_path + "ori_data/steel_ori_text.xlsx"
		self.train_file = self.data_path + "ner_data"

	def get_ner_data(self):
		excel=load_workbook(filename=self.train_file_ori,read_only=True)
		ws = excel.get_sheet_by_name('Sheet1')
		rows=ws.rows
		max_row=ws.max_row #获取行数
		#max_column=ws.max_column #获取列数
		max_column = 5
		print("转换任务开始，共{}行,{}列".format(max_row,max_column))
		data=list(rows)
		# 保存关键字
		keys = []
		# 保存结果
		result = []
		#读取标题作为key
		for x in range(0,max_column):
			keys.append(data[0][x].value)
		# 用于读取第一列和后面列1对多的excel文件
		for i in range(1,2022):
			recrod={}
			text_org=None
			for j in range(0,max_column):
				content = data[i][j].value
				if j == 0:
					recrod["id"] = content
				elif j == 1:
					if content!=None:
						content = self.pre_handle_text(str(content))
						text_org = content
						recrod["text"] = [i for i in content]
						recrod["labels"] = ["O"] * len(recrod['text'])
				elif j>1:
					if content!=None:
						content = str(content)
						if keys[j]=='specification':
							content = self.pre_handle_text(content)
						try:
							d = eval(content)
							find_start=0
							for rel_id, spo in enumerate(d):
								spo_start=0
								spo_end=0
								if spo.get("pos")!=None:
									spo_start = spo["pos"][0]
									spo_end = spo["pos"][1]
								else:
									spo_start = text_org.find(spo["name"],find_start)
									if spo_start == -1:
										continue
									spo_end = spo_start+len(spo["name"])-1
									find_start = spo_end+1
								recrod["labels"][spo_start] = "B-" + keys[j]
								if spo_end >spo_start:
									for q in range(spo_start + 1, spo_end+1):
									    recrod["labels"][q] = "I-" + keys[j]
						except :
							find_start=0
							for item in content.split():
								spo_start = text_org.find(item,find_start)
								if spo_start == -1:
									continue
								spo_end = spo_start+len(item)-1
								find_start = spo_end+1
								recrod["labels"][spo_start] = "B-" + keys[j]
								for q in range(spo_start + 1, spo_end+1):
									recrod["labels"][q] = "I-" + keys[j]
			result.append(recrod)
		train_ratio = 1
		train_num = int(len(result) * train_ratio)
		train_data = result[:train_num]
		dev_data = result[train_num:]
		#输出
		with open(self.train_file + "/train.txt", "w") as fp:
		    fp.write(os.linesep.join([json.dumps(d, ensure_ascii=False) for d in train_data]))

		with open(self.train_file + "/dev.txt", "w") as fp:
		    fp.write(os.linesep.join([json.dumps(d, ensure_ascii=False) for d in dev_data]))

		# 这里标签一般从数据中处理得到，这里我们自定义
		labels = keys[2:]
		with open(self.train_file + "/labels.txt", "w") as fp:
		    fp.write("\n".join(labels))
		print("转换任务结束")
	#文本纠错
	def pre_handle_text(self,text):
		text = re.sub(r'\n','<br />',text)
		text = re.sub(r'[xX]','*',text)
		text = re.sub(r' +',' ',text)
		text = re.sub(r' *\* *','*',text)
		text = re.sub(r' *(都是|/\*) *','*',text)
		text = re.sub(r' *或 *|？|\?|/+','/',text)
		return text

if __name__ == "__main__":
	processWineData = ProcessWineData()
	processWineData.get_ner_data()