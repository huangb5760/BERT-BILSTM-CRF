from openpyxl import load_workbook
import re
import json

# excel=load_workbook(filename='/Users/biao.huang/Downloads/公海客户20230717132557.xlsx',read_only=True)
# ws = excel.get_sheet_by_name('Sheet1')
# rows=ws.rows
# max_row=ws.max_row #获取行数
#max_column=ws.max_column #获取列数
def params_check(text):
    if not text:
        return "输入的文本为空"
    match_obj = re.search('\d{4}-\d{4}',text)
    if match_obj:
        return "价格不明确"
def optimize_business(text):
        if text == None:
            return
        result = []
        for id, item in enumerate(text):
            specification = item.get('specification')
            if specification:
                match_obj = re.match('\d+([件|吨])*',specification)
                if match_obj:
                	item['specification'] = match_obj.group(0)
                else:
                	item['specification'] = None
            result.append(item)
        return result
def has_number(text):
    match_obj = re.search('\d',text)
    if match_obj:
        return True
    return False
print(params_check('华磊 3670-3690'))
text = [{'specification':'3'},{'specification':'3件货'},{'specification':'3吨货'},{'specification':'3台货'},{'specification':'none'}]
print(optimize_business(text))
print(json.dumps(optimize_business(text)))
print(re.sub('[\u4e00-\u9fa5]', '', '给你3*5件货'))
print(has_number('识别1'))
