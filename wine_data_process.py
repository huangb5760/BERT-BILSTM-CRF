from openpyxl import load_workbook
import json
import os

class ProcessWineData:
	def __init__(self):
		self.data_path = "./data/wine/"
		self.train_file_ori = self.data_path + "ori_data/wine_org_text_gpt.xlsx"
		self.train_file = self.data_path + "ner_data"

	def get_ner_data(self):
		excel=load_workbook(filename=self.train_file_ori,read_only=True)
		ws = excel.get_sheet_by_name('Sheet1')
		rows=ws.rows
		max_row=ws.max_row #获取行数
		max_column=ws.max_column #获取列数
		print("转换任务开始，共{}行,{}列".format(max_row,max_column))
		data=list(rows)
		# 保存关键字
		keys = []
		# 保存结果
		result = []
		#读取标题作为key
		for x in range(0,max_column):
			keys.append(data[0][x].value)
		# 用于读取第一列和后面列1对多的excel文件
		for i in range(1,max_row):
			recrod={}
			text_org=None
			for j in range(0,max_column):
				content = data[i][j].value
				if j == 0:
					recrod["id"] = content
				elif j == 1:
					if content!=None:
						text_org = content
						recrod["text"] = [i for i in content]
						recrod["labels"] = ["O"] * len(recrod['text'])
				elif j>1:
					if content!=None:
						content = str(content)
						try:
							d = eval(content)
							find_start=0
							for rel_id, spo in enumerate(d):
								name = spo["name"]
								spo_start=0
								spo_end=0
								if spo.get("pos")!=None:
									spo_start = spo["pos"][0]
									spo_end = spo["pos"][1]
								else:
									spo_start = text_org.find(name,find_start)
									if spo_start == -1:
										continue
									spo_end = spo_start+len(name)-1
									find_start = spo_end+1
								recrod["labels"][spo_start] = "B-" + keys[j]
								if spo_end >spo_start:
									for q in range(spo_start + 1, spo_end+1):
									    recrod["labels"][q] = "I-" + keys[j]
						except :
							find_start=0
							for item in content.split('\n'):
								spo_start = text_org.find(item,find_start)
								if spo_start == -1:
									continue
								spo_end = spo_start+len(item)-1
								find_start = spo_end+1
								recrod["labels"][spo_start] = "B-" + keys[j]
								for q in range(spo_start + 1, spo_end+1):
									recrod["labels"][q] = "I-" + keys[j]
			result.append(recrod)
		train_ratio = 1
		train_num = int(len(result) * train_ratio)
		train_data = result[:train_num]
		dev_data = result[train_num:]
		#输出
		with open(self.train_file + "/train.txt", "w") as fp:
		    fp.write(os.linesep.join([json.dumps(d, ensure_ascii=False) for d in train_data]))

		with open(self.train_file + "/dev.txt", "w") as fp:
		    fp.write(os.linesep.join([json.dumps(d, ensure_ascii=False) for d in dev_data]))

		# 这里标签一般从数据中处理得到，这里我们自定义
		labels = keys[2:]
		with open(self.train_file + "/labels.txt", "w") as fp:
		    fp.write("\n".join(labels))
		print("转换任务结束")

	def generate_gpt_train_data(self):
		prompt_input = """我是一名白酒行业的数据分析专家,
可以从一段文字解析出sku(意向(sell/buy)、年份、品牌、系列、酒精度、规格、数量、价格)，按下面json输出格式
 ---
{'sku':[{'intention':string,// 意向:sell,buy
   'year':string, //年份
   'brand':string, //品牌
   'series':string, //系列
   'degree':string, //酒精度
   'specification':string, //规格
   'quantity':string,// 数量
   'price':string, //价格
}]}"""
		excel=load_workbook(filename=self.train_file_ori,read_only=True)
		ws = excel.get_sheet_by_name('Sheet1')
		rows=ws.rows
		max_row=ws.max_row #获取行数
		max_column=ws.max_column #获取列数
		print("转换任务开始，共{}行,{}列".format(max_row,max_column))
		data=list(rows)
		# 保存关键字
		keys = []
		# 保存结果
		result = []
		#读取标题作为key
		for x in range(0,max_column):
			keys.append(data[0][x].value)
		current_id = None
		text_org = None
		sku = []
		# 用于读取第一列和后面列1对多的excel文件
		for i in range(1,max_row):
			recrod={}
			if data[i][0].value!=None:
				if i>1:
					answer = {'sku':sku}
					unit = []
					unit.append({"role": "system", "content":prompt_input})
					unit.append({"role": "user", "content": text_org})
					unit.append({"role": "assistant", "content":json.dumps(answer, ensure_ascii=False)})
					message = {'messages':unit}
					result.append(json.dumps(message, ensure_ascii=False)) 
					sku = []
				current_id = data[i][0].value
				text_org=data[i][1].value
			for j in range(0,max_column):
				content = data[i][j].value
				if j>1 and content!=None:
					recrod[keys[j]] = str(content)
			sku.append(recrod)
			if i==max_row-1:
				answer = {'sku':sku}
				unit = []
				unit.append({"role": "system", "content":prompt_input})
				unit.append({"role": "user", "content": text_org})
				unit.append({"role": "assistant", "content":json.dumps(answer, ensure_ascii=False)})
				message = {'messages':unit}
				result.append(json.dumps(message, ensure_ascii=False)) 
		#输出
		with open(self.train_file + "/gpt_train.jsonl", "w") as fp:
		    fp.write(os.linesep.join(result))
		print("转换任务结束")

if __name__ == "__main__":
	processWineData = ProcessWineData()
	processWineData.generate_gpt_train_data()