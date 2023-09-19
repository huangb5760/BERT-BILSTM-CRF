from openpyxl import load_workbook
import json
import os

class ProcessWineData:
	def __init__(self):
		self.data_path = "./data/huodai/"
		self.train_file_ori = self.data_path + "ori_data/huodai_org_text_gpt.xlsx"
		self.train_file = self.data_path + "ner_data"
	def generate_gpt_train_data(self):
		prompt_input = """我是一名国际货物运输行业的数据分析专家,
可以从一段文字解析出场景、起始港、目的港、截单时间、截关时间、补料时间、离港时间、箱型、箱量、船司、价格，按下面json输出格式
 ---
[{
    'type'：string,//场景
    'startPort':string,//起始港
    'endPort':string,//目的港
    'cutDate':date,//截单时间
    'cyDate':date//截关时间
    'siDate':date//补料时间
    'edtDate':date//离港时间
    'boxType':string//箱型
    'boxCount':number//箱量
    'company':string//船司
    'price':number//价格
}]"""
		excel=load_workbook(filename=self.train_file_ori,read_only=True)
		ws = excel.get_sheet_by_name('Sheet1')
		rows=ws.rows
		max_row=ws.max_row #获取行数
		max_column=ws.max_column #获取列数
		print("转换任务开始，共{}行,{}列".format(max_row,max_column))
		data=list(rows)
		# 保存关键字
		keys = []
		# 保存结果
		result = []
		#读取标题作为key
		for x in range(0,max_column):
			keys.append(data[0][x].value)
		current_id = None
		text_org = None
		sku = []
		# 用于读取第一列和后面列1对多的excel文件
		for i in range(1,max_row):
			recrod={}
			if data[i][0].value!=None:
				if i>1:
					answer = {'sku':sku}
					unit = []
					unit.append({"role": "system", "content":prompt_input})
					unit.append({"role": "user", "content": text_org})
					unit.append({"role": "assistant", "content":json.dumps(answer, ensure_ascii=False)})
					message = {'messages':unit}
					result.append(json.dumps(message, ensure_ascii=False)) 
					sku = []
				current_id = data[i][0].value
				text_org=data[i][1].value
			for j in range(0,max_column):
				content = data[i][j].value
				if j>1 and content!=None:
					recrod[keys[j]] = str(content)
			sku.append(recrod)
			if i==max_row-1:
				unit = []
				unit.append({"role": "system", "content":prompt_input})
				unit.append({"role": "user", "content": text_org})
				unit.append({"role": "assistant", "content":json.dumps(sku, ensure_ascii=False)})
				message = {'messages':unit}
				result.append(json.dumps(message, ensure_ascii=False)) 
		#输出
		with open(self.train_file + "/gpt_train.jsonl", "w") as fp:
		    fp.write(os.linesep.join(result))
		print("转换任务结束")

if __name__ == "__main__":
	processWineData = ProcessWineData()
	processWineData.generate_gpt_train_data()